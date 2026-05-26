import csv
import os
import sys
import argparse
import unicodedata
from pathlib import Path

_PROJECT_ROOT = str(Path(__file__).resolve().parents[1])
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from backend.app.config.env import clean_env_value, get_supabase_client, load_env
from backend.app.embeddings.local_embeddings import embed_texts as embed_texts_local


def _iter_rows(path: Path) -> list[list[str]]:
    with path.open("rb") as fb:
        header = fb.read(4)

    if header.startswith(b"PK\x03\x04"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError(
                "El archivo parece ser un Excel (.xlsx) aunque tenga extensión .csv. "
                "Instala openpyxl (por ejemplo: py -m pip install openpyxl) o exporta el archivo a CSV real."
            ) from exc

        import tempfile

        with path.open("rb") as fb:
            raw = fb.read()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name

        try:
            wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if v is None else str(v) for v in row])
            return rows
        finally:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass

    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError:
        text = raw.decode("utf-8-sig", errors="replace")

    lines = text.splitlines()
    base_rows = list(csv.reader(lines, delimiter=",", quotechar='"', doublequote=True))
    if not base_rows:
        return []

    def normalize_row(row: list[str], expected_cols: int | None) -> list[str]:
        if len(row) == 1 and isinstance(row[0], str) and ("," in row[0]):
            try:
                nested = next(csv.reader([row[0]], delimiter=",", quotechar='"', doublequote=True))
                row = [str(x) for x in nested]
            except Exception:
                row = [str(x) for x in row]
        else:
            row = [str(x) for x in row]

        if expected_cols is not None and len(row) != expected_cols and row:
            joined = ",".join(row)
            try:
                reparsed = next(csv.reader([joined], delimiter=",", quotechar='"', doublequote=True))
                if len(reparsed) == expected_cols:
                    return [str(x) for x in reparsed]
            except Exception:
                pass

        return row

    header_row = normalize_row(base_rows[0], expected_cols=None)
    expected = len(header_row) if header_row else None

    normalized: list[list[str]] = [header_row]
    for row in base_rows[1:]:
        normalized.append(normalize_row(row, expected_cols=expected))

    return normalized


def _build_contenido(nombre: str, laboratorio: str, principio_activo: str, tipo_receta: str) -> str:
    return (
        f"El medicamento {nombre} tiene como principio activo {principio_activo}. "
        f"Es fabricado por el laboratorio {laboratorio}. "
        f"Condición de venta: {tipo_receta}."
    )


def _norm_key(text: str) -> str:
    raw = (text or "").strip().lower()
    return "".join(ch for ch in unicodedata.normalize("NFD", raw) if unicodedata.category(ch) != "Mn")


def _pick_idx(headers: list[str], candidates: list[str]) -> int | None:
    norm_headers = [_norm_key(h) for h in headers]
    for cand in candidates:
        c = _norm_key(cand)
        for idx, h in enumerate(norm_headers):
            if h == c:
                return idx
    for cand in candidates:
        c = _norm_key(cand)
        for idx, h in enumerate(norm_headers):
            if c and c in h:
                return idx
    return None


def _build_contenido_from_fields(nombre: str, fields: dict[str, str]) -> str:
    parts: list[str] = []
    for k, v in fields.items():
        v_clean = (v or "").strip()
        if not v_clean:
            continue
        parts.append(f"{k}: {v_clean}")
    joined = " | ".join(parts).strip()
    if joined:
        return f"{nombre}. {joined}"
    return nombre


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("ruta", nargs="?", default="data/raw/Medicamentos_.csv")
    parser.add_argument("--table", default="")
    parser.add_argument("--nivel-acceso", default="ambos")
    parser.add_argument("--col-nombre", default="")
    parser.add_argument("--col-laboratorio", default="")
    parser.add_argument("--col-principio-activo", default="")
    parser.add_argument("--col-tipo-receta", default="")
    parser.add_argument("--col-categoria", default="")
    parser.add_argument("--col-para-que-sirve", default="")
    parser.add_argument("--col-dosis-habitual", default="")
    parser.add_argument("--col-dosis-maxima", default="")
    parser.add_argument("--col-precauciones", default="")
    parser.add_argument("--col-efectos-secundarios", default="")
    parser.add_argument("--col-dieta-especial", default="")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--batch-size", type=int, default=50)
    args = parser.parse_args()

    ruta_archivo = args.ruta
    print(f"Abriendo archivo: {ruta_archivo}...")

    project_root = Path(__file__).resolve().parents[1]
    path = Path(ruta_archivo)
    if not path.is_absolute():
        path = (project_root / path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo: {path}")

    load_env()
    supabase = get_supabase_client()

    rows = _iter_rows(path)

    if not rows:
        print("CSV vacío")
        return 0

    headers = rows[0]
    data_rows = rows[1:] if len(rows) > 1 else []
    if args.offset:
        data_rows = data_rows[args.offset :]
    if args.limit and args.limit > 0:
        data_rows = data_rows[: args.limit]

    print(f"Se van a procesar {len(data_rows)} medicamentos.\n")

    table = (args.table or clean_env_value(os.getenv("SUPABASE_RAG_TABLE")) or "documentos_medicos").strip()

    idx_nombre = _pick_idx(headers, [args.col_nombre] if args.col_nombre else ["medicamento", "nombre", "nombre_medicamento"])
    idx_laboratorio = _pick_idx(headers, [args.col_laboratorio] if args.col_laboratorio else ["laboratorio", "fabricante", "marca"])
    idx_principio = _pick_idx(headers, [args.col_principio_activo] if args.col_principio_activo else ["principio activo", "principio_activo", "principio"])
    idx_tipo_receta = _pick_idx(headers, [args.col_tipo_receta] if args.col_tipo_receta else ["tipo receta", "tipo_receta", "condicion de venta", "condición de venta"])
    idx_categoria = _pick_idx(headers, [args.col_categoria] if args.col_categoria else ["categoria", "categoría", "clase"])
    idx_para = _pick_idx(headers, [args.col_para_que_sirve] if args.col_para_que_sirve else ["para que sirve", "para qué sirve", "indicaciones", "uso"])
    idx_dosis_habitual = _pick_idx(
        headers,
        [args.col_dosis_habitual]
        if args.col_dosis_habitual
        else ["dosis terapeutica", "dosis terapéutica", "dosis habitual", "dosis_habitual", "dosis"],
    )
    idx_dosis_maxima = _pick_idx(
        headers,
        [args.col_dosis_maxima] if args.col_dosis_maxima else ["dosis maxima", "dosis máxima", "dosis_maxima", "maximo", "máximo"],
    )
    idx_precauciones = _pick_idx(headers, [args.col_precauciones] if args.col_precauciones else ["precauciones", "advertencias", "contraindicaciones"])
    idx_efectos = _pick_idx(headers, [args.col_efectos_secundarios] if args.col_efectos_secundarios else ["efectos secundarios", "efectos", "reacciones adversas"])
    idx_dieta = _pick_idx(headers, [args.col_dieta_especial] if args.col_dieta_especial else ["dieta especial", "dieta", "alimentacion", "alimentación"])

    batch_items: list[dict[str, object]] = []
    batch_texts: list[str] = []
    inserted = 0

    def flush_batch() -> None:
        nonlocal inserted
        nonlocal batch_items
        nonlocal batch_texts
        if not batch_items:
            return
        vectors = embed_texts_local(batch_texts)
        payloads: list[dict[str, object]] = []
        for item, vec in zip(batch_items, vectors, strict=True):
            payloads.append(
                {
                    "nombre_medicamento": item["nombre_medicamento"],
                    "categoria": item.get("categoria"),
                    "para_que_sirve": item.get("para_que_sirve"),
                    "dosis_habitual": item.get("dosis_habitual"),
                    "dosis_maxima": item.get("dosis_maxima"),
                    "precauciones": item.get("precauciones"),
                    "efectos_secundarios": item.get("efectos_secundarios"),
                    "dieta_especial": item.get("dieta_especial"),
                    "contenido": item["contenido"],
                    "nivel_acceso": item["nivel_acceso"],
                    "embedding": vec,
                }
            )

        try:
            supabase.table(table).insert(payloads).execute()
        except Exception as exc:
            raise RuntimeError(
                f"Error insertando en Supabase (tabla={table}). Revisa SUPABASE_URL/SUPABASE_SERVICE_ROLE_KEY y tu conexión a internet. Detalle: {exc}"
            ) from exc
        inserted += len(payloads)
        batch_items = []
        batch_texts = []

    for index, row in enumerate(data_rows, start=1):
        try:
            if not row:
                raise ValueError("Fila vacía")

            if idx_nombre is None or idx_nombre >= len(row):
                raise ValueError("No se pudo identificar la columna de nombre/medicamento")

            if len(row) <= idx_nombre:
                raise ValueError("Fila con columnas insuficientes")

            nombre = str(row[idx_nombre]).strip()

            categoria = ""
            para_que_sirve = ""
            dosis_habitual = ""
            dosis_maxima = ""
            precauciones = ""
            efectos_secundarios = ""
            dieta_especial = ""

            fields: dict[str, str] = {}
            if idx_categoria is not None and idx_categoria < len(row):
                categoria = str(row[idx_categoria]).strip()
                fields["Categoría"] = categoria
            if idx_para is not None and idx_para < len(row):
                para_que_sirve = str(row[idx_para]).strip()
                fields["Para qué sirve"] = para_que_sirve
            if idx_dosis_habitual is not None and idx_dosis_habitual < len(row):
                dosis_habitual = str(row[idx_dosis_habitual]).strip()
                fields["Dosis habitual"] = dosis_habitual
            if idx_dosis_maxima is not None and idx_dosis_maxima < len(row):
                dosis_maxima = str(row[idx_dosis_maxima]).strip()
                fields["Dosis máxima"] = dosis_maxima
            if idx_precauciones is not None and idx_precauciones < len(row):
                precauciones = str(row[idx_precauciones]).strip()
                fields["Precauciones"] = precauciones
            if idx_efectos is not None and idx_efectos < len(row):
                efectos_secundarios = str(row[idx_efectos]).strip()
                fields["Efectos secundarios"] = efectos_secundarios
            if idx_dieta is not None and idx_dieta < len(row):
                dieta_especial = str(row[idx_dieta]).strip()
                fields["Dieta especial"] = dieta_especial

            if idx_laboratorio is not None and idx_laboratorio < len(row):
                fields["Laboratorio"] = str(row[idx_laboratorio]).strip()
            if idx_principio is not None and idx_principio < len(row):
                fields["Principio activo"] = str(row[idx_principio]).strip()
            if idx_tipo_receta is not None and idx_tipo_receta < len(row):
                fields["Condición de venta"] = str(row[idx_tipo_receta]).strip()

            if fields:
                texto_contenido = _build_contenido_from_fields(nombre, fields)
            else:
                texto_contenido = nombre

            print(f"Procesando: {nombre}...")

            datos_insertar = {
                "nombre_medicamento": nombre,
                "categoria": categoria or None,
                "para_que_sirve": para_que_sirve or None,
                "dosis_habitual": dosis_habitual or None,
                "dosis_maxima": dosis_maxima or None,
                "precauciones": precauciones or None,
                "efectos_secundarios": efectos_secundarios or None,
                "dieta_especial": dieta_especial or None,
                "contenido": texto_contenido,
                "nivel_acceso": str(args.nivel_acceso or "ambos"),
            }

            batch_items.append(datos_insertar)
            batch_texts.append(texto_contenido)
            if len(batch_items) >= args.batch_size:
                flush_batch()
                print(f"  ✅ Insertados: {inserted}")
            
        except Exception as e:
            print(f"  ❌ Error en la fila {index}: {e}")
            if "Error insertando en Supabase" in str(e):
                raise

    flush_batch()
    print(f"\n🎉 ¡Carga finalizada! Total insertados: {inserted}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
